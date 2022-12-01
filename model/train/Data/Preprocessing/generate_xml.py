import pandas as pd
import openpyxl as op
from openpyxl.drawing.image import Image

# 엑셀 파일 및 시트 객체 생성
wb = op.Workbook()
ws = wb.active

df = pd.read_csv("select.csv")
path = "../raw/raw_"

for j, row in df.iterrows():
    i = row['i']
    r = j + 2
    img = Image(path + str(i).zfill(5) + '.png')
    img.height = 200
    img.width = 200
    ws.add_image(img, 'A'+str(r))
    ws['B'+str(r)] = row['CHART']
    ws['C'+str(r)] = row['FORM_CODE_NAME']
    ws.row_dimensions[r].height = 100
ws.column_dimensions['A'].width = 100

#저장
wb.save(r"result.xlsx")